import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def creer_tableau_excel(nom_fichier="Calcul_Ventes_AE.xlsx"):
    """
    Crée un fichier Excel structuré pour la gestion des ventes AE.
    Structure :
      - Ligne 1 : Titre (B1:I1 fusionnées)
      - Ligne 3 : En-têtes du tableau
      - Lignes 4 à 19 : Données des composants avec formules et alternance de couleurs
      - Ligne 20 : Main d'œuvre (B20:D20 fusionnées pour le label, valeurs en E20, F20, calcul en I20)
      - Ligne 21 : Total (B21:D21 fusionnées pour le label, totaux calculés)
      - Lignes 23 à 25 : Section commission (colonnes D à I réinitialisées)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventes AE - Base"

    # --- Titre (ligne 1 : cellules B1:I1 fusionnées) ---
    ws.merge_cells("B1:I1")
    ws["B1"] = "Titre du Document"
    title_fill = PatternFill(start_color="8EAADB", fill_type="solid")
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    thick_border = Border(
        left=Side(border_style="thick", color="000000"),
        right=Side(border_style="thick", color="000000"),
        top=Side(border_style="thick", color="000000"),
        bottom=Side(border_style="thick", color="000000"),
    )
    for col in range(2, 10):
        cell = ws.cell(row=1, column=col)
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = title_alignment
        cell.border = thick_border

    # --- Largeur des colonnes ---
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20

    # --- Styles de base ---
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="B8CCE4", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", fill_type="solid")
    row_fill_even = PatternFill(start_color="F2F2F2", fill_type="solid")
    black_fill = PatternFill(start_color="000000", fill_type="solid")
    black_font = Font(color="000000")
    row23_fill = PatternFill(start_color="FFCC99", fill_type="solid")
    row24_fill = PatternFill(start_color="FFD966", fill_type="solid")
    row25_fill = PatternFill(start_color="FFEB9C", fill_type="solid")

    # --- En-têtes (ligne 3) ---
    headers = ["Composant", "Référence", "Lien", "Prix Unitaire TTC",
               "Quantité", "Total Fournisseur TTC", "Commission TTC", "Prix Client TTC"]
    for i, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=i)
        cell.value = header
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # --- Données (lignes 4 à 19) ---
    for row in range(4, 20):
        ws[f"E{row}"] = 0.0
        ws[f"F{row}"] = 0
        ws[f"G{row}"] = f"=E{row}*F{row}"
        ws[f"H{row}"] = f"=G{row}*$C$23"
        ws[f"I{row}"] = f"=G{row}+H{row}"
        ws[f"E{row}"].number_format = '#,##0.00 €'
        ws[f"G{row}"].number_format = '#,##0.00 €'
        ws[f"H{row}"].number_format = '#,##0.00 €'
        ws[f"I{row}"].number_format = '#,##0.00 €'
        ws[f"F{row}"].number_format = '0'
        fill = row_fill_even if row % 2 == 0 else row_fill_odd
        for col in range(2, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = bold_font
            cell.fill = fill

    # --- Main d'œuvre (ligne 20) ---
    ws.merge_cells("B20:D20")
    ws["B20"] = "Main d'oeuvre"
    ws["B20"].alignment = center_alignment
    ws["B20"].border = thin_border
    ws["B20"].font = bold_font
    ws["E20"] = 0.0
    ws["F20"] = 0
    ws["E20"].number_format = '#,##0.00 €'
    ws["F20"].number_format = '0'
    for col in ("E", "F"):
        cell = ws[f"{col}20"]
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.font = bold_font
    for col in ("G", "H"):
        cell = ws[f"{col}20"]
        cell.fill = black_fill
        cell.font = black_font
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.value = None
    ws["I20"] = "=E20*F20"
    ws["I20"].number_format = '#,##0.00 €'
    ws["I20"].border = thin_border
    ws["I20"].alignment = center_alignment
    ws["I20"].font = bold_font
    line20_fill = PatternFill(start_color="E2EFDA", fill_type="solid")
    for col in range(2, 10):
        if col in (7, 8):
            continue
        ws.cell(row=20, column=col).fill = line20_fill

    # --- Total (ligne 21) ---
    ws.merge_cells("B21:D21")
    ws["B21"] = "Total"
    ws["B21"].alignment = center_alignment
    ws["B21"].border = thin_border
    ws["B21"].font = bold_font
    for col in ("E", "F"):
        cell = ws[f"{col}21"]
        cell.fill = black_fill
        cell.font = black_font
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.value = None
    ws["G21"] = "=SUM(G4:G19)"
    ws["H21"] = "=SUM(H4:H19)"
    ws["I21"] = "=SUM(I4:I19)+I20"
    ws["G21"].number_format = '#,##0.00 €'
    ws["H21"].number_format = '#,##0.00 €'
    ws["I21"].number_format = '#,##0.00 €'
    line21_fill = PatternFill(start_color="BDD7EE", fill_type="solid")
    for col in range(2, 10):
        cell = ws.cell(row=21, column=col)
        cell.border = thin_border
        cell.fill = line21_fill
        cell.alignment = center_alignment
        cell.font = bold_font

    # --- Section Commission (lignes 23 à 25) ---
    ws["B23"] = "% Commission"
    ws["C23"] = 0.15
    ws["C23"].number_format = '0.00%'
    for col in range(2, 4):
        cell = ws.cell(row=23, column=col)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.font = bold_font
    for col in range(2, 10):
        ws.cell(row=23, column=col).fill = row23_fill
    for col in range(4, 10):
        cell = ws.cell(row=23, column=col)
        cell.border = None
        cell.fill = PatternFill()
    ws["B24"] = "Total Commission"
    ws["C24"] = "=SUM(H4:H19)"
    ws["C24"].number_format = '#,##0.00 €'
    ws["B25"] = "Total Commission + M.O."
    ws["C25"] = "=C24+I20"
    ws["C25"].number_format = '#,##0.00 €'
    for col in range(2, 10):
        cell24 = ws.cell(row=24, column=col)
        cell24.border = thin_border
        cell24.alignment = center_alignment
        cell24.font = bold_font
        cell24.fill = row24_fill if col < 4 else row_fill_odd
        cell25 = ws.cell(row=25, column=col)
        cell25.border = thin_border
        cell25.alignment = center_alignment
        cell25.font = bold_font
        cell25.fill = row25_fill if col < 4 else row_fill_odd
    for row in (24, 25):
        for col in range(4, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = None
            cell.fill = PatternFill()

    wb.save(nom_fichier)
    print(f"Le fichier '{nom_fichier}' a été créé avec succès.")

if __name__ == "__main__":
    creer_tableau_excel("Calcul_Ventes_AE.xlsx")
